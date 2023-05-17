import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import time
import logging

# ロギングの設定
logging.basicConfig(filename="output_percentage.log", level=logging.INFO)

# ファイルパス
# 開発用
file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/development.xlsx"  
# 本番用
# file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/ks_sideline-recommendation.xlsx"

try:
    # ワークブックを開く
    book = openpyxl.load_workbook(file_path)
    # シートを選択
    sheet = book['10位以内にランクインしているKW']

    # B列に値があれば削除
    for i in range(3, sheet.max_row + 1):
        if sheet.cell(row=i, column=2).value is not None:
            sheet.cell(row=i, column=2).value = None
        # B列が空白になったら処理を終了
        elif sheet.cell(row=i, column=2).value is None:
            break

    # 処理時間の計測開始
    start_time = time.time()

    # 各行のD列〜M列の間で10以下の数の割合を計算し、B列に出力
    for i in range(3, sheet.max_row + 1):
        values = [sheet.cell(row=i, column=j).value for j in range(4, 14)]
        count = sum(1 for v in values if isinstance(v, int) and v <= 10)
        ratio = count / 10
        sheet.cell(row=i, column=2).value = ratio

    # 処理時間の計測終了
    end_time = time.time()
    elapsed_time = end_time - start_time
    estimated_time = elapsed_time * (sheet.max_row - 2)

    # ログ出力
    logging.info(f"処理開始： {start_time}")
    logging.info(f"処理終了： {end_time}")
    logging.info(f"処理時間： {elapsed_time}秒")
    logging.info(f"推定処理時間： {estimated_time}秒")

    # ファイルの保存
    start_save = time.time()
    logging.info("保存開始...")
    book.save(file_path)
    end_save = time.time()
    logging.info("保存完了")
    logging.info(f"保存時間： {end_save - start_save}秒")

except InvalidFileException as e:
    print("無効なファイル形式です。")
    print("エラーメッセージ: ", str(e))
except Exception as e:
    print("予期せぬエラーが発生しました。")
    print("エラーメッセージ: ", str(e))
