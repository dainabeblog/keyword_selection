import openpyxl
import logging
import time

# ログ設定
logging.basicConfig(level=logging.INFO)

# 割合の閾値
ratio_threshold = 0.3  # この値を変更することで、出力されるキーワードの割合を制御できます

# ファイルパス
# 開発用
file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/development.xlsx"  
# 本番用
# file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/ks_sideline-recommendation.xlsx"  

# ワークブックを開く
book = openpyxl.load_workbook(file_path, keep_vba=True)

# ソースとターゲットのシートを取得
source_sheet = book['10位以内にランクインしているKW']
target_sheet = book['獲得すべきKW']

# ターゲットシートの2行目以降を空白にする
for row in range(2, target_sheet.max_row + 1):
    target_sheet.cell(row=row, column=1, value=None)
    target_sheet.cell(row=row, column=2, value=None)

# ターゲットシートの行インデックス
target_row = 2  # 2行目から開始

# ソースシートの行をループ
start_time = time.time()
total_rows = source_sheet.max_row - 3 + 1
for row in range(3, source_sheet.max_row + 1):  # 3行目から開始
    # A列が空白だった場合、それ以降のセルはスキップする
    keyword = source_sheet.cell(row=row, column=1).value  # A列
    if keyword is None or keyword == '':
        break

    # B列の値を取得
    cell_value = source_sheet.cell(row=row, column=2).value  # B列

    # B列の値を浮動小数点数に変換
    try:
        # B列の値が '%' を含む場合、それを削除してから数値に変換
        if isinstance(cell_value, str) and '%' in cell_value:
            cell_value = cell_value.replace('%', '')
            cell_value = float(cell_value) / 100
        else:
            cell_value = float(cell_value)
    except ValueError:
        logging.warning(f"B列の {row} 行目を浮動小数点数に変換できません。セルの値: '{cell_value}'. スキップします...")
        continue  # 変換に失敗した場合、このループをスキップ

    # 割合の閾値と比較
    if cell_value >= ratio_threshold:
        # A列とB列をターゲットシートにコピー
        target_sheet.cell(row=target_row, column=1, value=keyword)  # A列
        target_sheet.cell(row=target_row, column=2, value=source_sheet.cell(row=row, column=2).value)  # B列
        target_row += 1  # ターゲットシートの次の行に進む

    # 処理の進行状況を表示
    elapsed_time = time.time() - start_time
    progress = (row - 2) / total_rows  # -2 is because we start from the third row
    estimated_remaining_time = elapsed_time / progress * (1 - progress)
    logging.info(f"Processed row {row - 2} of {total_rows} ({progress * 100:.2f}%). "
                 f"Elapsed time: {elapsed_time:.2f}s. Estimated remaining time: {estimated_remaining_time:.2f}s.")

# ワークブックを保存
book.save(file_path)
logging.info("処理が正常に完了しました。")

