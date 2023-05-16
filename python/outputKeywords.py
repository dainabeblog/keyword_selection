import openpyxl
import logging

# ログ設定
logging.basicConfig(level=logging.INFO)

# ファイルパス
# 開発用
file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/development.xlsm"
# 本番用
# file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/ks_sideline-recommendation.xlsm"

# ワークブックを開く
book = openpyxl.load_workbook(file_path, keep_vba=True)

# ソースとターゲットのシートを取得
source_sheet = book['10位以内にランクインしているKW']
target_sheet = book['獲得すべきKW']

# ソースシートの行をループ
for row in range(3, source_sheet.max_row + 1):  # Start from 3rd row
    # Get the cell value
    cell_value = source_sheet.cell(row=row, column=2).value  # B column

    # Check if the cell value can be converted to a float
    try:
        # If the cell value includes '%', remove it and convert to decimal
        if isinstance(cell_value, str) and '%' in cell_value:
            cell_value = cell_value.replace('%', '')
            cell_value = float(cell_value) / 100
        else:
            cell_value = float(cell_value)
    except ValueError:
        logging.warning(f"Cannot convert cell B{row} to float. Skipping...")
        continue  # Skip this iteration if the conversion fails

    # Now we can safely do the comparison
    if cell_value >= 0.3:
        # Copy A and B columns to the target sheet
        target_sheet.cell(row=row - 1, column=1, value=source_sheet.cell(row=row, column=1).value)  # A column
        target_sheet.cell(row=row - 1, column=2, value=source_sheet.cell(row=row, column=2).value)  # B column

# Save the workbook
book.save(file_path)
logging.info("Process completed successfully.")
