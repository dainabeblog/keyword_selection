import pandas as pd
import openpyxl
import time
import logging
from tqdm import tqdm

# ログ設定
logging.basicConfig(level=logging.INFO)

start = time.time()

# ファイルパス
# 開発用
file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/development.xlsx"  # .xlsm -> .xlsx
# 本番用
# file_path = "/Users/watanabedaichi/Desktop/tomap/kw_selection/python/ks_sideline-recommendation.xlsx"  # .xlsm -> .xlsx

# ワークブックを開く
book = openpyxl.load_workbook(file_path)  # remove the 'keep_vba=True' option

# ターゲットのシートを取得
target_sheet = book['10位以内にランクインしているKW']

# 選択した範囲の列をループする
for column in tqdm(range(4, 14), desc="Processing columns"):
    sheet_number = target_sheet.cell(row=2, column=column).value
    logging.info(f"Processing column: {column}, Sheet number: {sheet_number}")
    if isinstance(sheet_number, int) and 1 <= sheet_number <= 10:
        # '1'から'10'までのシートを取得
        ws = book[str(sheet_number)]
        
        # キーワードと現在の位置の辞書を作成
        ws_dict = {}
        for row in range(2, ws.max_row + 1):  # Start from 2 assuming 1st row is header
            keyword = ws.cell(row=row, column=1).value  # Adjust column index as per your sheet
            curr_position = ws.cell(row=row, column=8).value  # Adjust column index as per your sheet
            ws_dict[keyword] = curr_position

        # ターゲットシートのA列についてループ
        for row in range(3, target_sheet.max_row + 1):  # Start from 3 assuming 1st and 2nd rows are header
            keyword = target_sheet.cell(row=row, column=1).value  # Keyword in column A
            if keyword is None or keyword == '':
                break  # If keyword is None or empty, stop processing this and the following rows
            elif keyword in ws_dict:
                # Keywordがある場合、対応する値を取得し、ターゲットシートに書き込む
                target_sheet.cell(row=row, column=column, value=ws_dict[keyword])  # Use cell method instead of at

        logging.info(f"Processed column: {column}")

# ワークブックを保存
book.save(file_path)

end = time.time()
logging.info(f"Execution time: {end - start} seconds")
